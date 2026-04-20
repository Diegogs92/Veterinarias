"""
Procesa el archivo de precios y genera un SQL de migración con categorías + productos.
Quita decimales y normaliza nombres.
"""
import openpyxl
import re
import uuid
import sys
from pathlib import Path

ROOT = Path(__file__).parent.parent
XLSX = ROOT / 'lista precis saludaniaml 821.xltx'
OUTPUT_SQL = ROOT / 'supabase' / 'migrations' / '20260420000002_seed_products.sql'

# ── Categorías ────────────────────────────────────────────────────────────────
CATEGORIES = [
    'Medicamentos',
    'Alimento balanceado',
    'Alimento húmedo',
    'Snacks y premios',
    'Higiene',
    'Accesorios',
    'Sanitario',
    'Platos y comederos',
]


def categorize(name: str) -> str:
    n = name.lower()
    if re.search(r'\bplato|comedero|mini\s+patitas|cara\s+de\s+gato', n):
        return 'Platos y comederos'
    if re.search(r'correa|collar|pretal|bolso|transporte|cardina|cepillo|chapita|mochila|manopla|funda|pañ|paño|donas|conjunto|pelota', n):
        return 'Accesorios'
    if re.search(r'piedra|bandeja|piedrita|pellcat|arena|sanitari', n):
        return 'Sanitario'
    if re.search(r'hueso|orejita|palito|galleta|bolsa\s+bon|bon\s*bon|golocan|golosin|premio', n):
        return 'Snacks y premios'
    if re.search(r'\d\s*kg|sieger|agility|eukanuba|royal|gooster|4\s+huellas|fit\s*32|fit\s*90|perrolac|gatolact', n):
        return 'Alimento balanceado'
    if re.search(r'\blata\b|salmon', n):
        return 'Alimento húmedo'
    if re.search(r'shampoo|shampo|cream|crema|talco|perfume|dental|limpia|curabicher|ecthol|dermosed|osspret|osspet|acederm|formula\s+mcdonald', n):
        return 'Higiene'
    return 'Medicamentos'


def clean_name(s) -> str:
    if s is None:
        return ''
    s = str(s).strip().replace('�', 'ñ')
    s = re.sub(r'\s+', ' ', s)
    # Cerrar paréntesis abiertos sin cerrar
    if s.count('(') > s.count(')'):
        s = s + ')'
    # Normalizar case: title case, pero preserva acrónimos cortos en mayúsculas
    words = s.split(' ')
    out = []
    for w in words:
        if not w:
            continue
        if w.isupper() and 2 < len(w) <= 5 and w.isalpha():
            out.append(w)  # acrónimos como MSD, NORT
        elif w.isupper() and len(w) <= 2:
            out.append(w)  # N1, S, etc
        else:
            # Title case respetando números/símbolos
            out.append(w[0].upper() + w[1:].lower() if len(w) > 1 else w.upper())
    result = ' '.join(out)
    # Limpiar paréntesis vacíos
    result = re.sub(r'\(\s*\)', '', result).strip()
    return result


def is_garbage(name: str) -> bool:
    """Filtrar nombres que son puro ruido: números solos, medidas, encabezados."""
    if not name:
        return True
    n = name.strip()
    # Una sola letra o número
    if len(n) <= 2:
        return True
    # Solo números/medidas (ej: "0.875", "3*12", "4/5", "5/6")
    if re.match(r'^[\d\s\.\,\*\/\+]+$', n):
        return True
    # Solo medidas tipo "16cm", "30 cm"
    if re.match(r'^\d+\s*cm$', n, re.I):
        return True
    # Rangos numéricos "1", "2", "3" en aislados
    if re.match(r'^N?\d+$', n, re.I):
        return True
    return False


def parse_price(v):
    if v is None or v == '':
        return None
    try:
        n = float(v)
        if n < 100:  # Precios menores a $100 probablemente errores
            return None
        return int(round(n))
    except (ValueError, TypeError):
        return None


def main():
    if not XLSX.exists():
        print(f'No se encontro {XLSX}')
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=False)
    products = []
    seen = set()

    def add(name, price, lab=None):
        if not name:
            return
        name = clean_name(name)
        if is_garbage(name):
            return
        p = parse_price(price)
        if p is None:
            return
        if lab:
            lab_c = clean_name(lab).strip()
            if lab_c and len(lab_c) > 1 and not lab_c.isdigit() and lab_c.lower() not in name.lower():
                name = f'{name} ({lab_c})'
        key = name.lower()
        if key in seen:
            return
        seen.add(key)
        products.append({
            'name': name,
            'price': p,
            'category': categorize(name),
        })

    # ── Hoja "Adelante" — productos principales ────────────────────────────
    ws = wb['Adelante']
    for row in range(2, 199):
        add(ws.cell(row=row, column=1).value,
            ws.cell(row=row, column=3).value,
            ws.cell(row=row, column=2).value)

    # ── Hoja "Adelante" — productos extra (filas 200+, 2 columnas) ─────────
    cat_left = None
    cat_right = None
    for row in range(200, 251):
        l_name = ws.cell(row=row, column=1).value
        l_price = ws.cell(row=row, column=2).value
        r_name = ws.cell(row=row, column=3).value
        r_price = ws.cell(row=row, column=4).value

        # Lado izquierdo
        if l_name and (l_price is None or parse_price(l_price) is None):
            # Encabezado de sección (ej: "CORREA LISA COMUN")
            clean = clean_name(l_name)
            if not is_garbage(clean):
                cat_left = clean
        elif l_name and l_price is not None:
            full = f'{cat_left} {l_name}' if cat_left else str(l_name)
            add(full, l_price)

        # Lado derecho
        if r_name and (r_price is None or parse_price(r_price) is None):
            clean = clean_name(r_name)
            if not is_garbage(clean):
                cat_right = clean
        elif r_name and r_price is not None:
            full = f'{cat_right} {r_name}' if cat_right else str(r_name)
            add(full, r_price)

    # ── Hoja "Alimento" ─────────────────────────────────────────────────────
    ws_al = wb['Alimento']
    for row in range(3, 90):
        add(ws_al.cell(row=row, column=1).value,
            ws_al.cell(row=row, column=3).value)

    products.sort(key=lambda p: (p['category'], p['name']))

    # ── Generar SQL ────────────────────────────────────────────────────────
    lines = [
        '-- ============================================================',
        '-- Seed: categorias + productos desde lista de precios',
        '-- ============================================================',
        '',
        '-- CATEGORIAS',
    ]

    cat_ids = {}
    for cat in CATEGORIES:
        cid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f'saludanimal.category.{cat}'))
        cat_ids[cat] = cid
        lines.append(
            f"INSERT INTO product_categories (id, name) VALUES ('{cid}', '{cat}') "
            f"ON CONFLICT (id) DO NOTHING;"
        )

    lines.extend(['', '-- PRODUCTOS'])
    for p in products:
        pid = str(uuid.uuid5(uuid.NAMESPACE_DNS, f'saludanimal.product.{p["name"].lower()}'))
        name_esc = p['name'].replace("'", "''")
        lines.append(
            f"INSERT INTO products (id, name, category_id, price, in_stock) VALUES "
            f"('{pid}', '{name_esc}', '{cat_ids[p['category']]}', {p['price']}, TRUE) "
            f"ON CONFLICT (id) DO NOTHING;"
        )

    OUTPUT_SQL.write_text('\n'.join(lines), encoding='utf-8')
    print(f'{len(products)} productos en {len(CATEGORIES)} categorias')
    print(f'SQL: {OUTPUT_SQL}')

    from collections import Counter
    counts = Counter(p['category'] for p in products)
    print('\nPor categoria:')
    for cat, n in counts.most_common():
        print(f'  {cat}: {n}')


if __name__ == '__main__':
    main()
